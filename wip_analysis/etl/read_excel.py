from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook

from wip_analysis.config import settings


SUMMARY_TOKENS = ("SUM", "TOTAL", "\u03a3", settings.u("\u5c0f\u8a08"), settings.u("\u5408\u8a08"))
FY_PATTERN = re.compile(r"\bFY\d{2}\b", re.IGNORECASE)


def normalize_header(value):
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\r\n", " ").replace("\n", " ").replace("\r", " ").replace("\t", " ")
    text = text.strip()
    return re.sub(r"\s+", " ", text)


def has_value(value):
    if value is None:
        return False
    if isinstance(value, str) and not value.strip():
        return False
    return True


def translate_header(header):
    if not header:
        return ""
    mapped = settings.FIELD_TRANSLATIONS.get(header)
    return mapped if mapped else header


def make_unique_headers(headers):
    counts = {}
    result = []
    for idx, header in enumerate(headers, start=1):
        base = header or f"column_{idx}"
        if base not in counts:
            counts[base] = 0
            result.append(base)
            continue
        counts[base] += 1
        result.append(f"{base}-{counts[base]}")
    return result


def build_headers(raw_headers, translate, sheet_name, mapping_entries):
    normalized = [normalize_header(value) for value in raw_headers]
    translated = [translate_header(value) if translate else value for value in normalized]
    unique_headers = make_unique_headers(translated)
    for original, final in zip(normalized, unique_headers):
        if not original:
            continue
        mapping_entries.append(
            {
                "source_sheet": sheet_name,
                "original_name": original,
                "translated_name": final,
            }
        )
    return unique_headers


def add_serialnumber(df, columns, serial_col="serialnumber"):
    if not columns or serial_col in df.columns:
        return df
    if not all(col in df.columns for col in columns):
        return df
    safe_df = df[columns].fillna("").astype(str)
    df[serial_col] = safe_df.agg("|".join, axis=1)
    return df


def forward_fill(values):
    filled = []
    last = ""
    for value in values:
        text = normalize_header(value)
        if text:
            last = text
            filled.append(text)
        else:
            filled.append(last)
    return filled


def is_summary_row(values):
    texts = [normalize_header(v) for v in values if isinstance(v, str)]
    combined = " ".join([t for t in texts if t])
    if not combined:
        return False
    upper = combined.upper()
    if FY_PATTERN.search(upper):
        return True
    for token in SUMMARY_TOKENS:
        if token and token in combined:
            return True
    return False


def read_simple_sheet(wb, sheet_name, header_row, data_start_row, translate, mapping_entries):
    ws = wb[sheet_name]
    max_col = ws.max_column
    header_cells = next(
        ws.iter_rows(min_row=header_row, max_row=header_row, min_col=1, max_col=max_col)
    )
    raw_headers = [cell.value for cell in header_cells]
    headers = build_headers(raw_headers, translate, sheet_name, mapping_entries)
    rows = []
    for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, min_col=1, max_col=max_col):
        values = [cell.value for cell in row]
        if not any(has_value(v) for v in values):
            continue
        rows.append(values)
    df = pd.DataFrame(rows, columns=headers)
    return df.dropna(axis=1, how="all")


def read_shipping_plan(wb, mapping_entries):
    sheet_name = settings.SHEET_NAMES["Shipping_Plan"]
    ws = wb[sheet_name]
    max_col = ws.max_column
    row1_cells = next(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=max_col))
    row2_cells = next(ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=max_col))
    row1_values = [cell.value for cell in row1_cells]
    row2_values = [cell.value for cell in row2_cells]
    row1_filled = forward_fill(row1_values)
    combined_raw = []
    for part1, part2 in zip(row1_filled, row2_values):
        text1 = normalize_header(part1)
        text2 = normalize_header(part2)
        combined = f"{text1}{text2}".strip()
        combined_raw.append(combined)
    headers = build_headers(combined_raw, True, sheet_name, mapping_entries)
    rows = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=max_col):
        values = [cell.value for cell in row]
        if not any(has_value(v) for v in values):
            continue
        rows.append(values)
    df = pd.DataFrame(rows, columns=headers)
    return df.dropna(axis=1, how="all")


def read_build_plan(wb, mapping_entries):
    sheet_name = settings.SHEET_NAMES["Build_Plan"]
    ws = wb[sheet_name]
    max_col = ws.max_column
    header_cells = next(ws.iter_rows(min_row=4, max_row=4, min_col=1, max_col=max_col))
    raw_headers = [cell.value for cell in header_cells]
    mapping_headers = build_headers(raw_headers[:3], True, sheet_name, mapping_entries)
    build_plan_headers = build_headers(raw_headers[3:], True, sheet_name, mapping_entries)

    mapping_rows = []
    build_rows = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=max_col):
        values = [cell.value for cell in row]
        left = values[:3]
        right = values[3:]
        if any(has_value(v) for v in left):
            mapping_rows.append(left)
        if any(has_value(v) for v in right):
            if is_summary_row(right):
                continue
            build_rows.append(right)

    mapping_df = pd.DataFrame(mapping_rows, columns=mapping_headers).dropna(axis=1, how="all")
    build_df = pd.DataFrame(build_rows, columns=build_plan_headers).dropna(axis=1, how="all")
    return mapping_df, build_df


def build_field_mapping(entries):
    if not entries:
        return pd.DataFrame(columns=["source_sheet", "original_name", "translated_name"])
    df = pd.DataFrame(entries)
    return df.drop_duplicates(subset=["source_sheet", "original_name", "translated_name"]).reset_index(
        drop=True
    )


def read_excel(path):
    tables = {}
    mapping_entries = []
    with open(path, "rb") as handle:
        wb = load_workbook(handle, read_only=True, data_only=True)
        tables["WIP_Raw_data"] = read_simple_sheet(
            wb,
            settings.SHEET_NAMES["WIP_Raw_data"],
            header_row=1,
            data_start_row=2,
            translate=True,
            mapping_entries=mapping_entries,
        )
        tables["WIP_Raw_DATE"] = read_simple_sheet(
            wb,
            settings.SHEET_NAMES["WIP_Raw_DATE"],
            header_row=1,
            data_start_row=2,
            translate=True,
            mapping_entries=mapping_entries,
        )
        tables["Complete_WIP"] = read_simple_sheet(
            wb,
            settings.SHEET_NAMES["Complete_WIP"],
            header_row=1,
            data_start_row=2,
            translate=True,
            mapping_entries=mapping_entries,
        )
        tables["Extra_Completed_WIP"] = read_simple_sheet(
            wb,
            settings.SHEET_NAMES["Extra_Completed_WIP"],
            header_row=1,
            data_start_row=2,
            translate=True,
            mapping_entries=mapping_entries,
        )
        tables["Shipping_Plan"] = read_shipping_plan(wb, mapping_entries)
        mapping_df, build_df = read_build_plan(wb, mapping_entries)
        tables["Mapping_Table"] = mapping_df
        tables["Build+Plan"] = build_df

    for table_name, df in list(tables.items()):
        serial_cols = settings.SERIALNUMBER_RULES.get(table_name)
        tables[table_name] = add_serialnumber(df, serial_cols)

    tables["Field_Mapping"] = build_field_mapping(mapping_entries)
    return tables


def find_excel_files(folder, pattern=None):
    folder = Path(folder)
    if not folder.exists():
        return []
    regex = re.compile(pattern or settings.EXCEL_PATTERN)
    return [p for p in folder.iterdir() if p.is_file() and regex.match(p.name)]
